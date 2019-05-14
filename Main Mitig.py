import openpyxl as op
from openpyxl.cell import Cell

wb = op.load_workbook(filename='resolved.xlsx')
sheet = wb.worksheets[2]
BART = wb.worksheets[1]
# sheet = wb.get_sheet_by_name('Sheet')

BART["AA1"] = 'Mitigated'
BART["AB1"] = 'Reason'
sheet["AA1"] = 'Mitigated'
sheet["AB1"] = 'Reason'

fusion = ['OBIA', 'Oracle Fusion']

fusion_pty = ["Fusion Procurement Operations",
              "Oracle Fusion Support Team",
              "WB Fusion Business Support",
              "WB Fusion OBIA Support",
              "Fusion KPMG",
              "WB Fusion Technical Support",
              "P & L Team"]

thirdparties = ["Assignment Group",
                "WHB-DigitalSupport",
                "WHB-3PTY-Oracle",
                "WHB-Digital Development",
                "WHB-3PTY-Comarch",
                "WHB-ELS-Windows 10 Desktop/VDI",
                "WHB-3PTY-Hitachi",
                "WHB-3PTY-Adactus",
                "WHB-3PTY-Celestra",
                "WHB-ELS-Windows 10 Laptop",
                "WHB-3PTY-WRS",
                "WHB-3PTY-Other",
                "WHB-3PTY-RedBox",
                "WHB-3PTY-Altius",
                "IS Security",
                "WHB-Digital Environments",
                "WHB-3PTY-Ceridian",
                "WHB-Cognizant STIBO",
                "WHB-ELS-DE-Windows 10 Laptop",
                "WHB-3PTY-ContentAndCodeOffice365",
                "WHB-ELS-Office-365",
                "Service Delivery",
                "WHB-3PTY-ECKOH",
                "WHB-3PTY-FIS",
                "WHB-3PTY-Whishworks",
                "WHB-3PTY-Yammer",
                "Nativ",
                "WHB-3PTY-ZScaler",
                "WHB-Cognizant CRM",
                "WHB-3PTY-Merkle",
                "WHB-3PTY-ContentCode",
                "WHB-3PTY-Freshways",
                "WHB-Architecture-Support",
                "Adactus",
                "WHB-3PTY- Ergo",
                "WHB-ELS-Office 2016 on Windows 7",
                "WHB-3PTY-CostaApplicationSupport",
                "Data Admin Team",
                "WHB-3PTY-K&N-KuehneAndNagel",
                "WHB-3PTY-Praesto",
                "WHB-3PTY-Kyocera",
                "WHB-3PTY-Signify",
                "WHB-Roastery-FacilityTeam",
                "WHB-3PTY-Steria",
                "WHB-3PTY-CSC",
                "WHB-3PTY-NetcallQMax",
                "WHB-3PTY-Pancentric",
                "WHB-3PTY-Graphite",
                "WHB-3PTY-Condeco",
                "WHB-3PTY-Magento",
                "WHB-3PTY-Identity Experts",
                "CST",
                "WHB-3PTY-Datacash",
                "WHB-3PTY-ONFIDO",
                "WHB-ProjectTeam",
                "WHB-3PTY-Grass Roots",
                "WHB-3PTY-Mastercard",
                "WHB-3PTY-Atos",
                "Kaba",
                "WHB-ELS-Office-365-Service",
                "WHB-3PTY-Rosslyn",
                "WHB-3PTY-Engage"]


for rowNum in range(2, BART.max_row):
    BARTparty = BART.cell(row=rowNum, column=12).value
    BARTconfig = BART.cell(row=rowNum, column=3).value  # can be changed to .lower

    if BART.cell(row=rowNum, column=4).value == 'Breached':

        if BARTconfig in fusion:                                        # working
            BART.cell(row=rowNum, column=27).value = 'Yes'
            BART.cell(row=rowNum, column=28).value = 'Fusion'

        if BARTparty in fusion_pty:                                     # working
            BART.cell(row=rowNum, column=27).value = 'Yes'
            BART.cell(row=rowNum, column=28).value = 'Fusion'


for rowNum in range(2, sheet.max_row):
    config = sheet.cell(row=rowNum, column=3).value  # can be changed to .lower
    party = sheet.cell(row=rowNum, column=12).value
    shortdisc = sheet.cell(row=rowNum, column=10).value
    parent = sheet.cell(row=rowNum, column=25).value
    close = sheet.cell(row=rowNum, column=24).value

    if sheet.cell(row=rowNum, column=4).value == 'Breached':

        if config in fusion:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Fusion'

        if party in fusion_pty:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Fusion'

        if party in thirdparties:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if 'internal' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Internal'

        if 'Internal' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Internal'

        if 'SIEM' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Internal'

        if 'damage' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Physical'

        if 'physical' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Physical'

        if 'SR 3' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if 'WRSID' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if 'WHIT-' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if 'COSTA-1' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if 'IN1900' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if 'HIGH MEMORY UTILIZATION' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Internal'

        if '#IM01' in shortdisc:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = '3rd Party'

        if parent not in [None, '']:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Child'

        if party == 'Fujitsu Virtual Admin':
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Customer tooling'

        if 'ServiceNow' in config:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'Customer tooling'
    try:
        if 'duplicate' in close:
            sheet.cell(row=rowNum, column=27).value = 'Yes'
            sheet.cell(row=rowNum, column=28).value = 'duplicate'
    except (AttributeError, TypeError):
        continue

wb.save('resolved.xlsx')
